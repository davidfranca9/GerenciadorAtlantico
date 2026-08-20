"""Importa dados/CIDADES E UF.xlsx para a tabela cidades do Postgres.

Uso: python -m app.import_cidades
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .database import Base, SessionLocal, engine
from .models import Cidade

XLSX_PATH = Path(__file__).resolve().parents[1] / "dados" / "CIDADES E UF.xlsx"


def _linhas_planilha():
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or not row[0]:
            continue
        nome = str(row[0]).strip()
        uf = str(row[1]).strip().upper() if len(row) > 1 and row[1] else ""
        ibge = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        if not nome or not uf or len(uf) != 2:
            continue
        yield nome, uf, ibge


def backfill_ibge_codes(db) -> int:
    """Preenche o codigo IBGE de cidades ja importadas sem esse dado."""
    if not XLSX_PATH.exists():
        return 0
    pendentes = db.query(Cidade).filter((Cidade.ibge == "") | (Cidade.ibge.is_(None))).count()
    if pendentes == 0:
        return 0

    ibge_por_nome_uf = {(nome, uf): ibge for nome, uf, ibge in _linhas_planilha() if ibge}
    atualizados = 0
    for cidade in db.query(Cidade).filter((Cidade.ibge == "") | (Cidade.ibge.is_(None))).all():
        ibge = ibge_por_nome_uf.get((cidade.nome, cidade.uf))
        if ibge:
            cidade.ibge = ibge
            atualizados += 1
    if atualizados:
        db.commit()
    return atualizados


def main():
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    try:
        if db.query(Cidade).count() > 0:
            atualizados = backfill_ibge_codes(db)
            print(f"Tabela cidades ja possui dados. {atualizados} codigo(s) IBGE preenchido(s).")
            return

        total = 0
        for nome, uf, ibge in _linhas_planilha():
            db.add(Cidade(nome=nome, uf=uf, ibge=ibge))
            total += 1
        db.commit()
        print(f"{total} cidades importadas com sucesso.")
    finally:
        db.close()


if __name__ == "__main__":
    main()
